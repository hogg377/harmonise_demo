# import pickle from turtle 
#from turtle import end_fill

import numpy as np


class MenuLog:
    def __init__(self, session_id) -> None:
        self.session_id = session_id
        #the log uses this so it knows which config the results are stored against
        self.config_run_order = []
 
        self.user_details = { #'name' : '',
                        #'date': '',
                        'name' : '',
                        'date' : '',
                        'participantnumber' : '',
                        'english' : '',
                        'vision': '',
                        'colour': '',
                        'age' : '',
                        'gender' : '',
                        'games' : ''
                        # consent is handled in paper formating following feedback from beta trials
                        # 'consent14': False,
                        #'final_consent': False
                        }
        self.post_test_questions = {}


    def save_responses(self, menu, menu_id, current_config='none'):
        """
        This function reads the menu_id and then saves the relevant content from the menu to the MenuLog's internal data structures
        """

        data = menu.get_input_data()

        if menu_id == 0:
            self.pickleLog('logs/' + self.session_id)

        if menu_id == 11:
            pass
        elif menu_id == 12:
            pass

        elif menu_id == 13:
            pass

        #Menu 14 is no longer included following feedback from beta trials
        elif menu_id == 14:
            #self.user_details['name'] = data['name']
            #self.user_details['date'] = data['date']
            #if this is called to store details then consent has been given
            #   default consent is falue and confirmation of consent is by clicking the "consent and continue" button which triggers this storage
            #self.user_details['consent14'] = True
            pass

        elif menu_id == 60:
            
            for entry in self.user_details:

                if entry not in data:
                    self.user_details[entry] = 'not answered'
                else:
                    self.user_details[entry] = data[entry]

            self.user_details['participantnumber'] = data['participantnumber']


            print('USER DETAILS: ', self.user_details)
            # if 'english' in data:
            #     self.user_details['english'] = data['english']
            # else:
            #     self.user_details['english'] = 'not answered'

            # if 'vision' in data:
            #     self.user_details['vision'] = data['vision']
            # else:
            #     self.user_details['vision'] = 'not answered'

            # if 'vision' in data:
            #     self.user_details['vision'] = data['vision']
            # else:
            #     self.user_details['vision'] = 'not answered'

            # if 'vision' in data:
            #     self.user_details['vision'] = data['vision']
            # else:
            #     self.user_details['vision'] = 'not answered'

            # if 'vision' in data:
            #     self.user_details['vision'] = data['vision']
            # else:
            #     self.user_details['vision'] = 'not answered'

            # self.user_details['colour'] = data['colour']
            # self.user_details['name60'] = data['name']
            # # self.user_details['age'] = data['age']
            # # self.user_details['gender'] = data['gender']
            # # self.user_details['games'] = data['games']
            # self.user_details['date60'] = data['date']
            pass

   

        elif  menu_id == 92:
            #create an entry for the current config if it doesn't exist
            if not current_config in self.post_test_questions:
                self.post_test_questions[current_config] = {'behaviour_perception': '', 'faultOrMal': ''}
                #store the order the configs were created (and assumed run in!)
                self.config_run_order.append(current_config)
            
           
            self.post_test_questions[current_config]['behaviour_perception'] = data['behaviour_perception']
            self.post_test_questions[current_config]['faultOrMal'] = data['faultOrMal']

        else:
            #there's no content to save
            pass

        return

    
    def pickleLog(self, file_name):

        #create the directory if it doesn't exist
        import os
        if not os.path.exists(file_name):
            os.makedirs(file_name)

        import pickle
        data = [self.session_id, self.user_details]
        fileo = open(f'{file_name}/user_details.pkl', 'wb')
        pickle.dump(data,fileo)
        fileo.close()

        data = [self.session_id, self.config_run_order, self.post_test_questions]
        fileo = open(f'{file_name}/post_test_responses.pkl', 'wb')
        pickle.dump(data,fileo)
        fileo.close()
        return

    def save_toexcel(self):

        import pickle
        import xlsxwriter
        import openpyxl

        # entry = '20220729T150923'

        with open(r"Results/"+self.session_id+"/user_details.pkl", "rb") as input_file:
            user_details = pickle.load(input_file)

        with open(r"Results/"+self.session_id+"/post_test_responses.pkl", "rb") as input_file:
            question_responses = pickle.load(input_file)

        # with open(r"Results/"+entry+"/config_fam_2_Active.pkl", "rb") as input_file:
        #   sim_data = pickle.load(input_file)


        # print(user_details)
        # print('\n\n', question_responses)
        # print(sim_data['user_log'])


        directory = 'Results/' + self.session_id + '/summary.xlsx'


        workbook = openpyxl.load_workbook('template.xlsx')

        sheet = workbook.active

        print(sheet["C11"].value)
        print(user_details)

        print('\n\n Question responses: ', question_responses)
        # sheet["C12"] = 'hello'

        # Write personal details

        # sheet["C3"] = user_details['name']
        sheet["C1"] = user_details[0]
        # sheet["C4"] = str(user_details[1]['birth'])
        # sheet["C5"] = user_details[1]['sex']
        # sheet["C6"] = user_details[1]['vision'][0][0]
        # sheet["C7"] = user_details[1]['colour'][0][0]

        sheet["C3"] = user_details[1]['participantnumber']
        sheet["C4"] = str(user_details[1]['english'])
        sheet["C5"] = str(user_details[1]['vision'])
        sheet["C6"] = str(user_details[1]['colour'])
        #self.user_details['name60'] = data['name']
        sheet["C7"] = user_details[1]['age']
        sheet["C8"] = str(user_details[1]['gender'])
        sheet["C9"] = user_details[1]['games']


        # Save ordered data block

        entry_index = 1

        row = 16
        col = 4

        tot_trials = 15

        for n in range(tot_trials):

            entry_name = 'config_exp_' + str(entry_index) + '_passive'


            if entry_name in question_responses[2]:

                print(entry_name, ' is in the dict')
                # behaviour scores
                sheet.cell(row = row, column=col).value = question_responses[2][entry_name]['behaviour_perception']

                sheet.cell(row = row, column=col + 1).value = question_responses[2][entry_name]['faultOrMal']
            else:
                sheet.cell(row = row, column=col).value = 'Not completed'

                sheet.cell(row = row, column=col + 1).value = 'Not completed'


            entry_name = 'config_exp_' + str(entry_index) + '_active'

            if entry_name in question_responses[2]:

                # behaviour scores
                sheet.cell(row = row + 20, column=col).value = question_responses[2][entry_name]['behaviour_perception']

                sheet.cell(row = row + 20, column=col + 1).value = question_responses[2][entry_name]['faultOrMal']
            else:
                sheet.cell(row = row + 20, column=col).value = 'Not completed'

                sheet.cell(row = row + 20, column=col + 1).value = 'Not completed'

            entry_index += 1
            row += 1



        # Save the trial order
        row = 16
        col = 7


        for i in range(len(question_responses[1])):

            sheet.cell(row = row, column=col).value = question_responses[1][i]
            row += 1

        workbook.save(filename=directory)



class AgentState:
    '''
    Records the key information about an agents state history
    The update(...) method should be called once for every tick the agent is "alive"
    id: the unique identifier for the agent
    time_created:   the simulaltion tick whent he agent was created
    time_destroyed: the simulaltion tick whent he agent was removed,
                    -1 (default) means it was alive when the simulation ended
    positions:      2d numpy array where the ith "row" is the xi,yi position of the agent at tick i since it was created
    tick:           not used.  Idea was to use it as an internal tracker to playback the agents positions (since the may be different to world time)
    '''
    def __init__(self, id) -> None:
        self.id = id
        self.state = {}
        #this will be used to "play back" the state
        #   it's essentially local time for the agent from when it was created
        #   and so potentially different from simulation time.
        self.tick = 0
        return


    def initialise(self, time, init_position):
        #going to append stuff during the update so useful to set something on the first row we can ignore later
        #init_time = np.empty(1)
        #init_time[:] = np.NaN
        #init_position = np.empty((1,2))
        #init_position[:] = np.NaN
        self.state = {
            'id' : self.id,
            'time_created' : time,
            #'times' : init_time,
            'time_destroyed': -1,
            'positions' : np.atleast_2d(init_position),
            'time' : np.array([time]),
            'empowerment' : np.array([-1])
        }
        self.tick = 0
        return


    def destroy(self, time):
        self.state['time_destroyed'] = time
        return


    def update(self, time, position, empowerment=-1):
        #self.agent_state['times'] = np.append(self.agent_state['times'], time)
        self.state['positions'] = np.append(self.state['positions'], np.atleast_2d(position), axis=0)
        self.state['time'] = np.append(self.state['time'], [time], axis=0)
        self.state['empowerment'] = np.append(self.state['empowerment'], [empowerment], axis=0)
        self.tick+=1
        return

class UserInputs:
    def __init__(self) -> None:
        self.input_at_t = {}
        self.events_at_t = {}
        return

    def initialise(self):
        self.input_at_t = {}
        self.events_at_t = {}


    def mouseclick_at_t_template(self):
        foo = {
            'type' : [],
            'screen_position' : [],
            'grid_position' : [],
            'realtime' : []
        }
        return foo


    def agentevent_at_t_template(self):
        foo = {
            'id' : [],
            'event': [],
            'grid_position' : []
        }
        return foo


    def addMouseClick(self, time, realtime, mouse_button, screen_pos, grid_pos):
        if time not in self.input_at_t.keys():
            self.input_at_t[time] = self.mouseclick_at_t_template()

        self.input_at_t[time]['type'].append(mouse_button)
        self.input_at_t[time]['screen_position'].append(screen_pos)
        self.input_at_t[time]['grid_position'].append(grid_pos)
        self.input_at_t[time]['realtime'].append(realtime)
        return


    def addAgentEvent(self, time, agent_id, grid_pos, event_type):
        if time not in self.events_at_t.keys():
            self.events_at_t[time] = self.agentevent_at_t_template()

        if event_type == 'destroy':
            agent_id *=-1
        self.events_at_t[time]['id'].append(agent_id)
        self.events_at_t[time]['grid_position'].append(grid_pos)
        self.events_at_t[time]['event'].append(event_type)
        return
    #end function



class SimLogger:
    def __init__(self) -> None:
        #logging lists
        
        self.user_log = UserInputs()
        self.user_log.initialise()
        #self.config = ''
        #self.session_id = ''
        self.meta = {}

        self.robot_pos = None
        self.faulty_id = None
        self.mal_id = None
        self.happiness = None
        self.coverage = None

    def initialise(self, session_id, config_name, seed, control_active):
        
        self.robot_pos = None
        self.faulty_id = None
        self.mal_id = None
        self.happiness = None
        self.coverage = None

        self.user_log = UserInputs()
        self.user_log.initialise()
        #self.config = config_name
        #self.session_id = session_id
        self.meta = {   'config_name' : config_name, 
                        'session_id' : session_id,
                        'sim_seed' : seed,
                        'start_time' : None, 
                        'end_time' : None, 
                        'control_active' : control_active
                        }

    def record_Step(self, agents, ):
        print("record step")
        # Save robot positions

    def record_coverage(self, coverage):

        # Save coverage data for whole trial
        self.coverage = coverage

    def save_agentId(self, ):
        print("save agent")

    def recordStartTime(self, time):
        self.meta['start_time'] = time
        return


    def recordEndTime(self, time):
        self.meta['end_time'] = time  
        return     


    def getList(self, type):
        if type=='dog':
            m_list = self.dog_logs
        elif type=='sheep':
            m_list = self.sheep_logs
        return m_list




    def logPopulationStates(self, type, population, time):
        m_list = self.getList(type)

        for agent in population.m_agents:
            #if the agent is a dog then it will have an empowerment value else set the default to -1 (sheep don't calculate empowerment)
            #TODO add empowerment to the sheep class (or agent class!) and set the default to -1 then all agents would have the agent.m_empowerment parameter but only dogs update it
            if type =='dog':
                empowerment_value = agent.m_empowerment
            else:
                empowerment_value = -1

            m_list[agent.m_id].update(time, agent.m_position, empowerment_value)
        return


    def logPopulations(self, populations, time):
        #size a numpy array to hold the agent ids and agent times
        n_agents = 0
        for pop in populations:
            n_agents+=len(pop.m_agents)

        #iniitalise the numpy arrays to hold the world state
        ids_at_t = np.zeros(n_agents)
        positions_at_t = np.zeros((n_agents,2))

        #itterate through each agent in each population storing their id and position in the master lists
        idx = 0
        for pop in populations:
            for agent in pop.m_agents:
                ids_at_t[idx] = agent.m_id
                positions_at_t[idx,:] = agent.m_position
                idx+=1

        #store the results
        self.world_at_t[time] = {'ids' : ids_at_t, 'positions' : positions_at_t}
        return


    def pickleLog(self, file_name):
        import pickle
        data = {
            'robot_pos' : self.robot_pos,
            'faulty_id' : self.faulty_id,
            'mal_id' : self.mal_id,
            'happiness' : self.happiness,
            'coverage' : self.coverage,
            'user_log' : self.user_log,
            'meta_data' : self.meta
        }
        fileo = open(f'{file_name}.pkl', 'wb')
        pickle.dump(data, fileo)
        fileo.close()
        return


