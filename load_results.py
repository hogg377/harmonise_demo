import pickle
import xlsxwriter
import openpyxl

entry = 'emer'

with open(r"Results/"+entry+"/user_details.pkl", "rb") as input_file:
	user_details = pickle.load(input_file)

with open(r"Results/"+entry+"/post_test_responses.pkl", "rb") as input_file:
	question_responses = pickle.load(input_file)

# with open(r"Results/"+entry+"/config_fam_2_Active.pkl", "rb") as input_file:
# 	sim_data = pickle.load(input_file)


for line in user_details:
	print('\n' + str(line))

print('Question response: \n\n')

print('Trial order: ', question_responses[1])

for line in question_responses[2]:
	print('\n' + line + ': ' + str(question_responses[2][line]))



directory = 'Results/' + entry


workbook = openpyxl.load_workbook('template.xlsx')

sheet = workbook.active

print(sheet["C11"].value)
sheet["C12"] = 'hello'

# Write personal details

# sheet["C3"] = user_details['name']
sheet["C1"] = user_details[0]
sheet["C4"] = str(user_details[1]['birth'])
sheet["C5"] = user_details[1]['sex']
sheet["C6"] = user_details[1]['vision'][0][0]
sheet["C7"] = user_details[1]['colour'][0][0]

row = 12
col = 2


for i in range(len(question_responses[1])):

	sheet.cell(row = row, column=col).value = question_responses[1][i]

	# behaviour scores
	sheet.cell(row = row, column=col + 1).value = question_responses[2][question_responses[1][i]]['behaviour_perception']

	sheet.cell(row = row, column=col + 2).value = question_responses[2][question_responses[1][i]]['faultOrMal']

	row += 1



workbook.save(filename="Results/summary.xlsx")



# Workbook() takes one, non-optional, argument
# which is the filename that we want to create.
# workbook = xlsxwriter.Workbook('hello.xlsx')
 
# # The workbook object is then used to add new
# # worksheet via the add_worksheet() method.
# worksheet = workbook.add_worksheet()
 
# # Use the worksheet object to write
# # data via the write() method.
# worksheet.write('A1', 'Hello..')
# worksheet.write('B1', 'Geeks')
# worksheet.write('C1', 'For')
# worksheet.write('D1', 'Geeks')
 
# # Finally, close the Excel file
# # via the close() method.
# workbook.close()